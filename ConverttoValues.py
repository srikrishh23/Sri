from openpyxl import load_workbook

def Testsheet1_updateformula():

    filepath = r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx'

    wb = load_workbook(filepath,data_only = True)
    wb_with_formula = load_workbook(filepath, data_only=False)

    wb.active = 3
    wb_with_formula.active = 3

    ws = wb.active
    ws1 = wb_with_formula.active


    print(ws.cell(row=6, column=2).value)
    ws1.cell(row =6,column=2).value = ws.cell(row=6, column=2).value
    print(ws1.cell(row=6, column=2).value)

    wb_with_formula.save(r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx')





def main():

    Testsheet1_updateformula()
    print("done")






if __name__ == "__main__":
    main()