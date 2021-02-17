from openpyxl import load_workbook
#to load existing workbook or xlsx file import load_workbook


def Testsheet1_formulato_value():
    filepath = r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx'
    # Create an Object for a particular workbook by passing the filepath
    wb_source = load_workbook(filepath,data_only=True)
    # set the actiive sheet name here 1st sheet in the workbook refers as 0 and below number 3 refers 4th worksheet
    wb_source.active = 3
    # call particcular active worksheet with below ws object
    ws = wb_source.active

    wb_dest = load_workbook(filepath, data_only=False)
    # set the actiive sheet name here 1st sheet in the workbook refers as 0 and below number 3 refers 4th worksheet
    wb_dest.active = 3
    # call particcular active worksheet with below ws object
    ws_dest = wb_dest.active

    value_list = []

    for row in ws['b6:b11']:
        for cell in row:
            print(cell.value)
            value_list.append(cell.value)
            print(cell.value)
    count = 0
    print(value_list)
    for row in ws_dest['b6:b11']:
        for cell in row:
            cell.value = value_list[count]
            count = count + 1
    print(value_list)

    wb_dest.save(r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx')




    print("Formula to value is done ")


def main():



    Testsheet1_formulato_value()




if __name__ == "__main__":
    main()