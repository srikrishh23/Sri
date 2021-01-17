from openpyxl import load_workbook
#to load existing workbook or xlsx file import load_workbook

#=VLOOKUP(A6,Overall!A5:I11,4,0)
#Entering file path of the existing file we are using 'r' to avoid syntax error  (unicode error) "unicodeescape"
def Testsheet1_updateformula():

    filepath = r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx'
    #Create an Object for a particular workbook by passing the filepath
    wb = load_workbook(filepath)
    #set the actiive sheet name here 1st sheet in the workbook refers as 0 and below number 3 refers 4th worksheet
    wb.active = 3
    #call particcular active worksheet with below ws object
    ws = wb.active

    #using for loop pass the cell range to be updated with value cell.value helps to update with formula

    #formulas in each cell refers the lookupvalue works only with tuple / exact cell reference

    for row in ws['b6:b11']:
        for cell in row:
         cell.value = "=VLOOKUP(A{0}, 'Overall'!$A$5: $I$11, 3, FALSE)".format(cell.row)
    print("success B cell Ref updated with formula ")

    for row in ws['c6:c11']:
        for cell in row:
            cell.value = "=VLOOKUP(A{0}, 'Overall'!$A$5: $I$11, 4, FALSE)".format(cell.row)
    print("success B cell Ref updated with formula")

    wb.save(r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx')


def Testsheet2_updateformula():

    print("Now at Testsheet2")

    filepath = r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx'
    #Create an Object for a particular sheet
    wb = load_workbook(filepath)
    wb.active = 4
    ws = wb.active

    for row in ws['b6:b11']:
        for cell in row:
         cell.value = "=VLOOKUP(A{0}, 'Overall'!$A$5: $I$11, 3, FALSE)".format(cell.row)
    print("success B cell Ref updated with formula ")

    for row in ws['c6:c11']:
        for cell in row:
            cell.value = "=VLOOKUP(A{0}, 'Overall'!$A$5: $I$11, 4, FALSE)".format(cell.row)
    print("success B cell Ref updated with formula")

    wb.save(r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx')

def Testsheet3_updateformula():



    filepath = r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx'
    #Create an Object for a particular sheet
    wb = load_workbook(filepath)
    wb.active = 5
    ws = wb.active

    for row in ws['b6:b11']:
        for cell in row:
         cell.value = "=VLOOKUP(A{0}, 'Overall'!$A$5: $I$11, 3, FALSE)".format(cell.row)
    print("success B cell Ref updated with formula ")

    for row in ws['c6:c11']:
        for cell in row:
            cell.value = "=VLOOKUP(A{0}, 'Overall'!$A$5: $I$11, 4, FALSE)".format(cell.row)
    print("success B cell Ref updated with formula")

    wb.save(r'C:\Users\User\Desktop\Openpyxl tests\cc bills_Test.xlsx')

def main():

    Testsheet1_updateformula()
    print("Now at Testsheet2")
    Testsheet2_updateformula()
    print("Testsheet2 done")
    print("Now at Testsheet3")
    Testsheet3_updateformula()
    print("Testsheet3 done")



if __name__ == "__main__":
    main()